# Copyright (c) Instacertify
"""Upload helpers for Quote Format Library and Laboratory Scope Library."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

import frappe
from frappe import _
from frappe.utils import cint
from frappe.utils.file_manager import get_file

from instacertify.utils.files import assert_internal_file

_TAG_RE = re.compile(r"\[Tags:\s*([^\]]*)\]", re.IGNORECASE)


def _parse_tags(notes: str | None) -> list[str]:
	"""Extract comma-separated tags stored as [Tags: …] in template_notes."""
	if not notes:
		return []
	m = _TAG_RE.search(notes)
	if not m:
		return []
	return [t.strip() for t in (m.group(1) or "").split(",") if t.strip()]


def _apply_tags_to_notes(notes: str | None, tags: str) -> str:
	tag_line = f"[Tags: {tags.strip()}]"
	notes = notes or ""
	if _TAG_RE.search(notes):
		return _TAG_RE.sub(tag_line, notes).strip()
	return (notes + "\n" + tag_line).strip() if notes else tag_line


def _guess_file_type(filename: str | None) -> str:
	name = (filename or "").lower()
	if name.endswith(".pdf"):
		return "PDF"
	if name.endswith((".doc", ".docx")):
		return "DOCX"
	if name.endswith((".htm", ".html")):
		return "HTML"
	if name.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")):
		return "Image"
	if name.endswith((".xls", ".xlsx", ".csv")):
		return "Excel/CSV"
	return "Other"


@frappe.whitelist()
def create_quote_format_from_upload(
	template_name: str,
	quotation_type: str,
	file_url: str | None = None,
	service_family: str | None = None,
	template_notes: str | None = None,
	is_active: int = 1,
):
	"""Create (or update) an IC Quotation Template and attach an uploaded quote format."""
	template_name = (template_name or "").strip()
	quotation_type = (quotation_type or "").strip() or "Consulting"
	if not template_name:
		frappe.throw(_("Template Name is required"))
	if quotation_type not in (
		"Consulting",
		"Testing",
		"Renewal",
		"Other",
		"Multiple Products / Multiple Services",
		"Service",
	):
		frappe.throw(_("Invalid quotation type: {0}").format(quotation_type))

	exists = frappe.db.exists("IC Quotation Template", template_name)
	if exists:
		doc = frappe.get_doc("IC Quotation Template", template_name)
	else:
		doc = frappe.new_doc("IC Quotation Template")
		doc.template_name = template_name

	doc.quotation_type = quotation_type
	if service_family:
		doc.service_family = service_family
	if template_notes:
		doc.template_notes = template_notes
	doc.is_active = 1 if int(is_active or 1) else 0

	if file_url:
		file_url = assert_internal_file(file_url, "Quote format file")
		doc.uploaded_format = file_url
		label = Path(str(file_url).split("/")[-1] or "Quote Format").name
		# Avoid duplicate rows for same file
		already = any((r.attach_file or "") == file_url for r in (doc.format_library or []))
		if not already:
			doc.append(
				"format_library",
				{
					"file_label": label,
					"attach_file": file_url,
					"file_type": _guess_file_type(label),
					"remarks": "Uploaded via Quote Format Library",
				},
			)

	doc.save(ignore_permissions=False)
	frappe.db.commit()
	return {"template": doc.name, "uploaded_format": doc.uploaded_format}


@frappe.whitelist()
def create_laboratory_from_upload(
	laboratory_name: str,
	location: str | None = None,
	accreditation_scope: str | None = None,
	accreditation_details: str | None = None,
	scope_file: str | None = None,
	contact_person: str | None = None,
	email: str | None = None,
	phone: str | None = None,
	city: str | None = None,
	address: str | None = None,
	website: str | None = None,
	status: str = "Active",
	import_scopes_from_file: int = 1,
):
	"""Create/update a laboratory with editable master fields + optional scope file.

	If the attached scope file is CSV/Excel, its rows are imported into test_scopes
	(unless import_scopes_from_file is 0).
	"""
	laboratory_name = (laboratory_name or "").strip()
	if not laboratory_name:
		frappe.throw(_("Laboratory Name is required"))

	existing = frappe.db.get_value("IC Laboratory", {"laboratory_name": laboratory_name}, "name")
	if existing:
		doc = frappe.get_doc("IC Laboratory", existing)
	else:
		doc = frappe.new_doc("IC Laboratory")
		doc.laboratory_name = laboratory_name

	doc.status = status or "Active"
	# Always apply dialog values so edits stick on the form
	doc.location = (location or "").strip()
	doc.city = (city or "").strip()
	doc.address = (address or "").strip()
	doc.website = (website or "").strip()
	doc.contact_person = (contact_person or "").strip()
	doc.email = (email or "").strip()
	doc.phone = (phone or "").strip()
	if accreditation_scope is not None:
		doc.accreditation_scope = _plain_text(accreditation_scope)
	if accreditation_details is not None:
		doc.accreditation_details = _plain_text(accreditation_details)

	imported = 0
	updated = 0
	if scope_file:
		scope_file = assert_internal_file(scope_file, "Laboratory scope file")
		doc.scope_sheet = scope_file
		name_l = str(scope_file).lower()
		if name_l.endswith(".pdf"):
			doc.accreditation_scope_pdf = scope_file
		if int(import_scopes_from_file or 0) and name_l.endswith(
			(".csv", ".xlsx", ".xlsm", ".xls")
		):
			imported, updated = _apply_scope_rows_to_lab(doc, scope_file)

	doc.save(ignore_permissions=False)
	if scope_file:
		_link_file_to_lab(scope_file, doc.name, "scope_sheet")
	frappe.db.commit()
	return {
		"laboratory": doc.name,
		"laboratory_name": doc.laboratory_name,
		"scopes_imported": imported,
		"scopes_updated": updated,
		"scope_rows": len(doc.test_scopes or []),
	}


def _plain_text(value: str | None) -> str:
	"""Strip accidental HTML wrappers from Text Editor leftovers."""
	if not value:
		return ""
	text = str(value)
	# Collapse simple paragraph wrappers from older Text Editor values when re-saving
	if text.startswith("<") and ">" in text:
		from frappe.utils import strip_html

		text = strip_html(text)
	return text.strip()


def _link_file_to_lab(file_url: str, lab_name: str, fieldname: str):
	"""Point uploaded File rows at the laboratory so Attach fields stay editable."""
	try:
		names = frappe.get_all("File", filters={"file_url": file_url}, pluck="name")
		for name in names:
			frappe.db.set_value(
				"File",
				name,
				{
					"attached_to_doctype": "IC Laboratory",
					"attached_to_name": lab_name,
					"attached_to_field": fieldname,
				},
				update_modified=False,
			)
	except Exception:
		frappe.log_error(frappe.get_traceback(), "link lab scope file")


def _parse_money(value) -> float:
	if value in (None, ""):
		return 0.0
	if isinstance(value, (int, float)):
		return float(value)
	text = str(value).strip().replace(",", "").replace("₹", "").replace("INR", "").strip()
	if not text:
		return 0.0
	try:
		return float(text)
	except Exception:
		return 0.0


def _scope_row_key(test_name: str, standard: str) -> str:
	return f"{(test_name or '').strip().casefold()}||{(standard or '').strip().casefold()}"


def _apply_scope_rows_to_lab(doc, file_url: str) -> tuple[int, int]:
	"""Upsert test_scopes from CSV/Excel. Returns (added, updated)."""
	rows = _read_lab_scope_rows(file_url)
	if not rows:
		return 0, 0

	existing = {}
	for row in doc.get("test_scopes") or []:
		existing[_scope_row_key(row.test_name, row.applicable_standard)] = row

	added = 0
	updated = 0
	for item in rows:
		test_name = item.get("test_name") or ""
		if not test_name:
			continue
		standard = item.get("applicable_standard") or ""
		key = _scope_row_key(test_name, standard)
		payload = {
			"test_name": test_name,
			"applicable_standard": standard,
			"category": item.get("category") or "",
			"selling_price": _parse_money(item.get("selling_price")),
			"purchase_price": _parse_money(item.get("purchase_price")),
			"currency": item.get("currency") or "INR",
			"is_active": item.get("is_active", 1),
		}
		payload["margin"] = float(payload["selling_price"] or 0) - float(payload["purchase_price"] or 0)
		if key in existing:
			row = existing[key]
			for k, v in payload.items():
				row.set(k, v)
			updated += 1
		else:
			doc.append("test_scopes", payload)
			added += 1
	return added, updated


def _read_lab_scope_rows(file_url: str) -> list[dict]:
	"""Parse lab scope CSV/Excel into normalized dict rows."""
	content, file_name = _file_bytes_from_url(file_url)
	name_l = (file_name or file_url or "").lower()
	raw_rows: list[dict] = []

	if name_l.endswith((".xlsx", ".xlsm", ".xls")):
		from openpyxl import load_workbook

		wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
		ws = None
		for title in ("Scopes", "Scope", "Lab Scopes", "Sheet1"):
			if title in wb.sheetnames:
				ws = wb[title]
				break
		ws = ws or wb.active
		raw_rows = _rows_from_worksheet(ws)
	else:
		text = content.decode("utf-8-sig", errors="ignore")
		try:
			dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;")
		except Exception:
			dialect = csv.excel
		reader = csv.DictReader(io.StringIO(text), dialect=dialect)
		for raw in reader:
			row = _normalize_upload_row(raw)
			if any(row.values()):
				raw_rows.append(row)

	out = []
	for row in raw_rows:
		test_name = _pick(row, "test_name", "test", "name", "scope", "test name")
		if not test_name:
			continue
		active_raw = _pick(row, "is_active", "active")
		is_active = 1
		if active_raw != "":
			is_active = 1 if str(active_raw).strip().lower() in ("1", "true", "yes", "y", "active") else 0
		out.append(
			{
				"test_name": test_name,
				"applicable_standard": _pick(
					row, "applicable_standard", "standard", "applicable standard"
				),
				"category": _pick(row, "category", "type"),
				"selling_price": _pick(
					row, "selling_price", "price", "selling price", "rate"
				),
				"purchase_price": _pick(
					row, "purchase_price", "buying_price", "cost", "buying price"
				),
				"currency": _pick(row, "currency") or "INR",
				"is_active": is_active,
			}
		)
	return out


def _pick(row: dict, *keys) -> str:
	lower = {(k or "").strip().lower(): v for k, v in (row or {}).items()}
	for key in keys:
		if key in lower and lower[key] not in (None, ""):
			return str(lower[key]).strip()
	return ""


@frappe.whitelist()
def import_laboratory_scopes_csv(laboratory: str, file_url: str):
	"""Import / upsert test scope rows from CSV or Excel into a laboratory."""
	if not laboratory or not frappe.db.exists("IC Laboratory", laboratory):
		frappe.throw(_("Laboratory not found"))
	if not file_url:
		frappe.throw(_("Select a CSV or Excel file from My Device or File Library first"))

	file_url = assert_internal_file(file_url, "Scope spreadsheet")
	doc = frappe.get_doc("IC Laboratory", laboratory)
	added, updated = _apply_scope_rows_to_lab(doc, file_url)
	if not added and not updated:
		frappe.throw(
			_(
				"No scope rows found. Use headers like test_name, applicable_standard, category, selling_price, purchase_price."
			)
		)

	# Keep the latest spreadsheet on the lab for reference
	doc.scope_sheet = file_url
	doc.save(ignore_permissions=False)
	_link_file_to_lab(file_url, doc.name, "scope_sheet")
	frappe.db.commit()
	return {
		"laboratory": doc.name,
		"added": added,
		"updated": updated,
		"scope_rows": len(doc.test_scopes or []),
	}


@frappe.whitelist()
def get_library_summary():
	"""Counts for Home / explore prompts."""
	return {
		"quote_templates": frappe.db.count("IC Quotation Template", {"is_active": 1}),
		"laboratories": frappe.db.count("IC Laboratory", {"status": "Active"}),
		"labs_with_scope_file": frappe.db.count(
			"IC Laboratory",
			{"status": "Active", "scope_sheet": ["is", "set"]},
		),
		"templates_with_upload": frappe.db.sql(
			"""
			select count(*) from `tabIC Quotation Template`
			where ifnull(uploaded_format,'') != '' or ifnull(uploaded_template_pack,'') != ''
			"""
		)[0][0],
	}


@frappe.whitelist()
def get_quote_library_catalog():
	"""Category + tag catalog for Quote Format Library page."""
	rows = frappe.get_all(
		"IC Quotation Template",
		fields=[
			"name",
			"template_name",
			"quotation_type",
			"service_family",
			"service_name",
			"is_active",
			"uploaded_format",
			"template_notes",
			"modified",
		],
		order_by="quotation_type asc, template_name asc",
		limit_page_length=500,
	)
	# Default amounts per template (sales-editable on the template form)
	totals: dict[str, float] = {}
	line_counts: dict[str, int] = {}
	passthrough_counts: dict[str, int] = {}
	if rows:
		names = [r.name for r in rows]
		cost_rows = frappe.get_all(
			"IC Quotation Cost Item",
			filters={"parent": ["in", names], "parenttype": "IC Quotation Template"},
			fields=["parent", "amount", "is_passthrough", "revenue_treatment"],
			limit_page_length=5000,
		)
		for c in cost_rows:
			parent = c.parent
			line_counts[parent] = line_counts.get(parent, 0) + 1
			totals[parent] = totals.get(parent, 0.0) + float(c.amount or 0)
			treatment = (c.get("revenue_treatment") or "").strip()
			is_pass = treatment == "Do Not Count as Revenue" or cint(c.is_passthrough)
			if is_pass:
				passthrough_counts[parent] = passthrough_counts.get(parent, 0) + 1

	counts = {}
	for r in rows:
		t = r.quotation_type or "Other"
		counts[t] = counts.get(t, 0) + 1
		r["tags"] = _parse_tags(r.get("template_notes"))
		r["cost_line_count"] = line_counts.get(r.name, 0)
		r["default_amount_total"] = totals.get(r.name, 0.0)
		r["passthrough_line_count"] = passthrough_counts.get(r.name, 0)
	return {"counts": counts, "templates": rows}


def _public_file(file_name: str, content: str | bytes, content_type: str | None = None) -> dict:
	"""Create or reuse a public File and return its URL."""
	existing = frappe.db.get_value("File", {"file_name": file_name, "is_private": 0}, "file_url")
	if existing:
		return {"ok": True, "file_url": existing, "file_name": file_name}
	doc = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": file_name,
			"content": content,
			"is_private": 0,
		}
	)
	if content_type:
		doc.content_type = content_type
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"ok": True, "file_url": doc.file_url, "file_name": file_name}


@frappe.whitelist()
def download_lab_scope_template(fmt: str | None = None) -> dict:
	"""Downloadable CSV or Excel template for laboratory scope row import."""
	fmt = (fmt or "csv").strip().lower()
	headers = [
		"test_name",
		"applicable_standard",
		"category",
		"selling_price",
		"purchase_price",
		"currency",
		"is_active",
	]
	samples = [
		["EMI/EMC Radiated Emission", "CISPR 32", "EMC", 25000, 18000, "INR", 1],
		["Safety Insulation Resistance", "IEC 62368-1", "Safety", 8000, 5500, "INR", 1],
		["RF Conducted Spurious", "ETSI EN 300 328", "RF", 15000, 11000, "INR", 1],
	]
	if fmt in ("xls", "xlsx", "excel"):
		from openpyxl import Workbook
		from openpyxl.styles import Font

		wb = Workbook()
		ws = wb.active
		ws.title = "Scopes"
		ws.append(headers)
		for cell in ws[1]:
			cell.font = Font(bold=True)
		for row in samples:
			ws.append(row)
		guide = wb.create_sheet("Instructions", 0)
		guide["A1"] = "Instacertify Laboratory Scope Upload"
		guide["A1"].font = Font(bold=True, size=14)
		guide["A3"] = "1. Fill one row per accredited test on the Scopes sheet."
		guide["A4"] = "2. Upload via Laboratory → Upload Lab / Scope, or Import Scope CSV/Excel on the form."
		guide["A5"] = "3. Matching test_name + applicable_standard updates an existing row; new names are added."
		guide["A6"] = "4. is_active: 1 = Active, 0 = Inactive"
		out = io.BytesIO()
		wb.save(out)
		name = "IC_Laboratory_Scope_Upload_Template.xlsx"
		existing = frappe.db.get_value("File", {"file_name": name, "is_private": 0}, "name")
		if existing:
			frappe.delete_doc("File", existing, ignore_permissions=True, force=True)
		return _public_file(
			name,
			out.getvalue(),
			"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)

	buf = io.StringIO()
	writer = csv.writer(buf)
	writer.writerow(headers)
	writer.writerows(samples)
	content = ("\ufeff" + buf.getvalue()).encode("utf-8")
	name = "IC_Laboratory_Scope_Upload_Template.csv"
	existing = frappe.db.get_value("File", {"file_name": name, "is_private": 0}, "name")
	if existing:
		frappe.delete_doc("File", existing, ignore_permissions=True, force=True)
	return _public_file(name, content, "text/csv")


_QUOTE_TEMPLATE_HEADERS = [
	"template_name",
	"quotation_type",
	"service_family",
	"service_name",
	"certification_type",
	"applicable_standard",
	"estimated_timeline",
	"validity_days",
	"is_active",
	"template_notes",
	"tags",
]

_QUOTE_TEMPLATE_SAMPLES = [
	[
		"BIS CRS Consulting Pack",
		"Consulting",
		"BIS CRS",
		"BIS CRS Consultancy",
		"BIS CRS",
		"IS 13252",
		"4–6 weeks",
		"90",
		"1",
		"Standard BIS CRS consulting narrative",
		"BIS,CRS,Consulting",
	],
	[
		"EMI/EMC Testing Quote",
		"Testing",
		"EMC",
		"EMI/EMC Lab Testing",
		"EMC",
		"CISPR 32",
		"2–3 weeks",
		"60",
		"1",
		"Lab testing commercials from library",
		"EMC,Testing,Lab",
	],
	[
		"BIS Licence Renewal",
		"Renewal",
		"BIS Renewal",
		"BIS Licence Renewal",
		"BIS Renewal",
		"",
		"3–4 weeks",
		"90",
		"1",
		"Renewal pack",
		"BIS,Renewal",
	],
]


def _quote_template_csv_bytes() -> bytes:
	buf = io.StringIO()
	writer = csv.writer(buf)
	writer.writerow(_QUOTE_TEMPLATE_HEADERS)
	writer.writerows(_QUOTE_TEMPLATE_SAMPLES)
	# UTF-8 BOM helps Excel open CSV cleanly
	return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _quote_template_xlsx_bytes() -> bytes:
	from openpyxl import Workbook
	from openpyxl.styles import Font

	wb = Workbook()
	ws = wb.active
	ws.title = "Quote Templates"
	ws.append(_QUOTE_TEMPLATE_HEADERS)
	for cell in ws[1]:
		cell.font = Font(bold=True)
	for row in _QUOTE_TEMPLATE_SAMPLES:
		ws.append(row)
	guide = wb.create_sheet("Instructions", 0)
	guide["A1"] = "Instacertify Quote Format Library — Excel / CSV upload"
	guide["A1"].font = Font(bold=True, size=14)
	guide["A3"] = "1. Fill one row per format on the Quote Templates sheet."
	guide["A4"] = "2. quotation_type must be exactly one of:"
	guide["A5"] = "   Consulting, Testing, Renewal, Service, Other, Multiple Products / Multiple Services"
	guide["A6"] = "3. tags: comma-separated labels shown on the library page (e.g. BIS,CRS,Consulting)"
	guide["A7"] = "4. is_active: 1 = Active, 0 = Inactive"
	guide["A8"] = "5. Download Excel or CSV from Quote Format Library, fill rows, then Import spreadsheet."
	guide["A9"] = "6. Matching template_name updates an existing format; new names create formats."
	out = io.BytesIO()
	wb.save(out)
	return out.getvalue()


@frappe.whitelist()
def download_quote_format_upload_template(fmt: str | None = None) -> dict:
	"""Downloadable CSV or Excel template for Quote Format Library bulk upload (not HTML)."""
	fmt = (fmt or "xlsx").strip().lower()
	if fmt in ("xls", "xlsx", "excel"):
		content = _quote_template_xlsx_bytes()
		name = "IC_Quote_Format_Library_Template.xlsx"
		existing = frappe.db.get_value("File", {"file_name": name, "is_private": 0}, "name")
		if existing:
			frappe.delete_doc("File", existing, ignore_permissions=True, force=True)
		return _public_file(
			name,
			content,
			"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)

	content = _quote_template_csv_bytes()
	name = "IC_Quote_Format_Library_Template.csv"
	existing = frappe.db.get_value("File", {"file_name": name, "is_private": 0}, "name")
	if existing:
		frappe.delete_doc("File", existing, ignore_permissions=True, force=True)
	return _public_file(name, content, "text/csv")


def _file_bytes_from_url(file_url: str) -> tuple[bytes, str]:
	"""Return (bytes, file_name) for an internal File URL."""
	file_url = assert_internal_file(file_url, "Quote library spreadsheet")
	file_name = Path(str(file_url).split("?")[0]).name
	content = None
	try:
		file_doc = frappe.get_doc("File", {"file_url": file_url})
		file_name = file_doc.file_name or file_name
		content = file_doc.get_content()
	except Exception:
		content = None
	if content is None:
		# get_file returns [file_name, content]
		got = get_file(file_url)
		if isinstance(got, (list, tuple)) and len(got) >= 2:
			file_name = got[0] or file_name
			content = got[1]
		else:
			content = got
	if content is None:
		frappe.throw(_("Could not read uploaded file"))
	if isinstance(content, str):
		content = content.encode("utf-8")
	return content, file_name


def _cell_str(val) -> str:
	"""Normalize a CSV/Excel cell to a stripped string (handles overflow lists)."""
	if val is None:
		return ""
	if isinstance(val, (list, tuple)):
		return ",".join(_cell_str(x) for x in val if x is not None and _cell_str(x))
	return str(val).strip()


def _normalize_upload_row(raw: dict) -> dict:
	"""Lower-case keys; fold DictReader overflow (None key) into tags."""
	row: dict[str, str] = {}
	overflow: list[str] = []
	for k, v in (raw or {}).items():
		key = (k or "").strip().lower() if isinstance(k, str) else k
		val = _cell_str(v)
		if key in (None, ""):
			if val:
				overflow.extend([p.strip() for p in val.split(",") if p.strip()])
			continue
		row[key] = val
	if overflow:
		extra = ",".join(overflow)
		row["tags"] = f"{row['tags']},{extra}" if row.get("tags") else extra
	return row


def _rows_from_worksheet(ws) -> list[dict]:
	headers = None
	rows: list[dict] = []
	for i, row in enumerate(ws.iter_rows(values_only=True)):
		vals = [_cell_str(v) for v in row]
		if i == 0:
			headers = [h.lower().strip() for h in vals]
			continue
		if not any(vals) or not headers:
			continue
		# Extra columns beyond headers → append to tags
		item = {headers[j]: vals[j] for j in range(min(len(headers), len(vals))) if headers[j]}
		if len(vals) > len(headers):
			extra = ",".join(v for v in vals[len(headers) :] if v)
			if extra:
				item["tags"] = f"{item.get('tags','')},{extra}" if item.get("tags") else extra
		rows.append(item)
	return rows


def _read_upload_rows(file_url: str) -> list[dict]:
	"""Parse CSV or Excel rows from an uploaded File URL."""
	content, file_name = _file_bytes_from_url(file_url)
	name_l = (file_name or "").lower()

	if name_l.endswith((".xlsx", ".xlsm", ".xls")):
		from openpyxl import load_workbook

		wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
		ws = None
		for title in ("Quote Templates", "Quote Formats", "Sheet1"):
			if title in wb.sheetnames:
				ws = wb[title]
				break
		if ws is None:
			for sheet in wb.worksheets:
				probe = _rows_from_worksheet(sheet)
				if probe and "template_name" in probe[0]:
					return probe
			ws = wb.active
		return _rows_from_worksheet(ws)

	text = content.decode("utf-8-sig")
	# Prefer excel dialect; fall back if sniff fails
	try:
		dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;")
	except Exception:
		dialect = csv.excel
	reader = csv.DictReader(io.StringIO(text), dialect=dialect)
	rows: list[dict] = []
	for raw in reader:
		row = _normalize_upload_row(raw)
		if any(row.values()):
			rows.append(row)
	return rows


@frappe.whitelist()
def import_quote_templates_from_spreadsheet(file_url: str):
	"""Bulk create/update IC Quotation Templates from CSV or Excel."""
	rows = _read_upload_rows(file_url)
	if not rows:
		frappe.throw(_("No data rows found in the spreadsheet"))

	created, updated, skipped = [], [], []
	valid_types = {
		"Consulting",
		"Testing",
		"Renewal",
		"Other",
		"Multiple Products / Multiple Services",
		"Service",
	}

	for row in rows:
		name = (row.get("template_name") or "").strip()
		qtype = (row.get("quotation_type") or "Consulting").strip() or "Consulting"
		if not name:
			skipped.append({"reason": "missing template_name", "row": row})
			continue
		if qtype not in valid_types:
			skipped.append({"reason": f"invalid quotation_type: {qtype}", "row": row})
			continue

		exists = frappe.db.exists("IC Quotation Template", name)
		doc = frappe.get_doc("IC Quotation Template", name) if exists else frappe.new_doc("IC Quotation Template")
		if not exists:
			doc.template_name = name

		doc.quotation_type = qtype
		for field in (
			"service_family",
			"service_name",
			"certification_type",
			"applicable_standard",
			"estimated_timeline",
			"template_notes",
		):
			if row.get(field):
				setattr(doc, field, row.get(field))
		if row.get("validity_days"):
			try:
				doc.validity_days = int(float(row["validity_days"]))
			except Exception:
				pass
		if "is_active" in row and row.get("is_active") != "":
			doc.is_active = 1 if str(row.get("is_active")).strip().lower() in ("1", "true", "yes", "y") else 0
		else:
			doc.is_active = 1

		tags = (row.get("tags") or "").strip()
		if tags:
			doc.template_notes = _apply_tags_to_notes(doc.template_notes, tags)

		doc.save(ignore_permissions=False)
		(updated if exists else created).append(doc.name)

	frappe.db.commit()
	return {
		"created": created,
		"updated": updated,
		"skipped": skipped,
		"created_count": len(created),
		"updated_count": len(updated),
		"skipped_count": len(skipped),
		"message": _(
			"Import finished: {0} created, {1} updated{2}"
		).format(
			len(created),
			len(updated),
			f", {len(skipped)} skipped" if skipped else "",
		),
	}
